import matplotlib.pyplot as plt
import numpy as np
import os
import glob
import pandas as pd
import pickle
import re
import string
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from scipy.sparse import csr_matrix

from age import age
from gender import gender
from year_level import year_level
from sections import sections
from sentiment_results import sentiments_bar, sentiments_percent
from generate_wordcloud import generate_wordcloud

script_directory = os.path.dirname(os.path.abspath(__file__))
script_directory = os.path.join(script_directory, "..", "..", "..", "uploads")

# read the data file
# change the path to the current file and save the row number of the feedback we need to predict
# path = file path
current_directory = os.path.dirname(os.path.abspath(__file__))
# path = os.path.join(script_directory, 'student_feedback.xlsx')
path = os.path.join('./public/uploads/student_feedback.xlsx')

# path = "result.xlsx"

dataset = pd.read_excel(path)  # pathing to the file
# dataset = pd.read_excel(path)  # pathing to the file
# remove N/A sentiments and repeated responses
lemmatizer = WordNetLemmatizer()

dataset.dropna(axis=0, inplace=True, subset=[
               "Comment/Suggestion for Improvement"])
# cleaning dataset and tokenize


def clean_text(txt):
    # remove punctuation
    txt_nopunct = "".join([x for x in txt if x not in string.punctuation])
    txt_nopunct = txt_nopunct.lower()  # lower casing all
    txt_token = re.split("\W", txt_nopunct)  # tokenization
    # remove stop words
    txt_stop = ([x for x in txt_token if x not in stopwords.words("english")])
    txt_stop = [x for x in txt_stop if x != ""]
    txt_lemmatize = []
    for word in txt_stop:
        txt_lemmatize.append(lemmatizer.lemmatize(word)
                             )  # lemmatizing each word
    # removing redundunt words in one statement
    txt_final = list(set(txt_lemmatize))
    return txt_final


# load trained models
saved_model = pickle.load(
    open(os.path.join('./public/models/StackedModel.pkl'), "rb"))
saved_vector = pickle.load(
    open(os.path.join('./public/models/Vectorizer.pkl'), "rb"))

# saved_model = pickle.load(
#     open(os.path.join(script_directory, "StackedModel.pkl"), "rb"))
# saved_vector = pickle.load(
#     open(os.path.join(script_directory, "Vectorizer.pkl"), "rb"))

# get the text from excel that needs prediction and its row number
dataset["Sentiment"] = np.nan
needs_pred_num = int(dataset["Sentiment"].isnull().sum())
needs_pred = dataset["Feedback"][-needs_pred_num:]

# Replace np.nan values with empty strings
needs_pred.fillna("", inplace=True)

# Do the prediction
vector_word = csr_matrix(saved_vector.transform(needs_pred))
new_pred = saved_model.predict(vector_word)

if needs_pred_num > 0:
    for row_num in range(needs_pred_num):
        current_row = row_num-needs_pred_num
        dataset["Sentiment"].iloc[current_row] = int(new_pred[current_row])

    # save the model_prediction to the row,col of the sentiment
    from openpyxl import load_workbook
    excel_editor = load_workbook(path)
    sheet = excel_editor.active
    for ws in excel_editor.worksheets:
        max_row = ws.max_row
        start_num = max_row - (max_row-2)
        am_del = max_row - start_num + 1
        ws.delete_rows(idx=start_num, amount=(am_del))
        col_names = [cell.value for cell in ws[1]]
        if "Sentiment" not in col_names:
            ws.cell(row=1, column=len(col_names)+1).value = "Sentiment"

        total_row = len(dataset["Sentiment"])
        for row_num in range(total_row):
            current_row = row_num - total_row
            ws.append(list(dataset.iloc[current_row]))

    # CSV ung taas tapos EXCEL ung baba tanggalin mo nalang kung ano di mo kailangan
    # tentative name
    df = pd.DataFrame(vector_word.toarray(),
                      columns=saved_vector.get_feature_names_out())

    dataset.to_csv("sentiment_analysis.csv", index=False, header=True)
    excel_editor.save(path)
    
    df = pd.read_csv("sentiment_analysis.csv")
    age(df)
    gender(df)
    sections(df)
    year_level(df)
    sentiments_bar(df)
    sentiments_percent(df)
    generate_wordcloud(df)
    print('TADAAAA')


else:

    import tkinter as tk
    from tkinter import ttk
    popup = tk.Tk()
    popup.wm_title("ERROR!")
    popup.geometry("170x100")
    popup.attributes("-toolwindow", True)
    label = ttk.Label(popup, text=" No New Row ",
                      font=("Verdana", 15), justify="center")
    label.pack(side="top", fill="x", pady=20)
    B1 = ttk.Button(popup, text="Okay", command=popup.destroy)
    B1.pack()
    popup.mainloop()
